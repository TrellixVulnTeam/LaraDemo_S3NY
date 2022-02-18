import openpyxl

workbook= openpyxl.load_workbook('/var/www/storage/app/uploads/Input.xlsx')
sh = workbook['input']

#print(sh.max_row)
#print(sh.max_column)
rows = sh.rows
Acase=[]
for row in list(rows):
    case=[]
    for c in row:
        case.append(c.value)
    Acase.append(case)

day=47
temp=47
gz=92
sheet = workbook['TEST']

for j in range(1,sh.max_column):
    num=1
    for i in range(0,sh.max_row):
        #print(num)
        #print(str(j)+":"+str(i))
        if Acase[i][j]!="" and Acase[i][j]!=None and num<20:
            str1="A"+str(day)
            str2="C"+str(day)
            sheet[str1]=str(Acase[i][0])
            sheet[str2]=str(Acase[i][j])
            day=day+2
            #print(str(day)+":"+Acase[i][0]+":"+Acase[i][j])
            num=num+1
    temp=temp+gz
    day=temp

workbook.save('/var/www/storage/app/public/output.xlsx')
